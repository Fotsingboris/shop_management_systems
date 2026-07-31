"""Vues du catalogue produit (EF-2 : catégories)."""
from __future__ import annotations

from typing import Any, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import QuerySet
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import CreateView, ListView, TemplateView, UpdateView

from products.forms import CategorieForm, CategorieImportForm
from products.models import Categorie, ImportCategories
from products.tasks import importer_categories


class AdminRequiredMixin(UserPassesTestMixin):
    """Réservé à l'Admin (EF-2.1 : gestion des catégories)."""

    def test_func(self) -> bool:
        return getattr(self.request.user, "is_admin", False)


class CategorieListView(LoginRequiredMixin, ListView):
    """Liste paginée des catégories, avec recherche (EF-2.1, EF-13.3)."""

    model = Categorie
    template_name = "dashboard/pages/admin/products/category/categorie_list.html"
    context_object_name = "categories"
    paginate_by = 20

    def get_queryset(self) -> QuerySet[Categorie]:
        # select_related("parent") : évite une requête par ligne pour
        # afficher le nom de la catégorie parente (EF-13.1).
        qs = Categorie.objects.select_related("parent").order_by("nom")
        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(nom__icontains=q)
        return qs

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "")
        return context


class CategorieCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    """Création d'une catégorie (EF-2.1)."""

    model = Categorie
    form_class = CategorieForm
    template_name = "dashboard/pages/admin/products/category/categorie_form.html"
    success_url = reverse_lazy("products:categorie_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _("Catégorie « %(nom)s » créée.") % {"nom": self.object.nom})
        return response


class CategorieUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    """Modification d'une catégorie existante (EF-2.1).

    Réutilise volontairement ``CategorieForm`` et le même template que la
    création (``categorie_form.html``) : le formulaire est identique, seul
    le titre/texte du bouton changent selon que ``form.instance.pk`` existe
    déjà ou non. C'est la vue qui manquait — le lien « Modifier » de
    ``categorie_list.html`` pointait vers ``#`` faute de cible.
    """

    model = Categorie
    form_class = CategorieForm
    template_name = "dashboard/pages/admin/products/category/categorie_form.html"
    success_url = reverse_lazy("products:categorie_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _("Catégorie « %(nom)s » mise à jour.") % {"nom": self.object.nom})
        return response


class CategorieImportView(LoginRequiredMixin, AdminRequiredMixin, TemplateView):
    """Upload d'un fichier Excel de catégories, traité en arrière-plan (EF-2.3).

    La requête ne fait que sauvegarder le fichier et déclencher la tâche
    Celery (``importer_categories.delay``) ; elle ne bloque donc jamais,
    même pour un fichier de plusieurs milliers de lignes. Le rapport de
    validation apparaît sur cette même page une fois la tâche terminée.
    """

    template_name = "dashboard/pages/admin/products/category/categorie_import.html"

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.setdefault("form", CategorieImportForm())
        context["imports"] = (
            ImportCategories.objects.select_related("importe_par").order_by("-created_at")[:10]
        )
        return context

    def post(self, request, *args, **kwargs):
        form = CategorieImportForm(request.POST, request.FILES)
        if form.is_valid():
            import_obj = form.save(commit=False)
            if request.user.is_authenticated:
                import_obj.importe_par = request.user
            import_obj.save()
            importer_categories.delay(str(import_obj.pk))
            messages.success(
                request,
                _("Import lancé. Le rapport apparaîtra ci-dessous une fois le traitement terminé."),
            )
            return redirect("products:categorie_import")

        context = self.get_context_data(form=form)
        return self.render_to_response(context)


class CategorieImportTemplateView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Génère et sert le modèle Excel attendu par l'import (EF-2.3).

    Généré à la volée plutôt que servi comme fichier statique : les
    colonnes restent forcément synchronisées avec ce que
    ``products.tasks.importer_categories`` lit réellement (nom, parent,
    actif), sans risque qu'un fichier statique parte en décalage avec le
    code au fil du temps.
    """

    FILENAME = "modele_import_categories.xlsx"
    HEADERS = ["nom", "parent", "actif"]
    EXEMPLES = [
        ["Alimentation", "", 1],
        ["Boissons", "Alimentation", 1],
        ["Boissons gazeuses", "Boissons", 1],
        ["Hygiène (désactivée)", "", 0],
    ]

    def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Catégories"

        sheet.append(self.HEADERS)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        for col_idx in range(1, len(self.HEADERS) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for row in self.EXEMPLES:
            sheet.append(row)

        for col_idx, width in enumerate([30, 30, 10], start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.FILENAME}"'
        workbook.save(response)
        return response