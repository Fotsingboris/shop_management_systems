"""Vues du catalogue produit (EF-3 : produits).

Fichier séparé de ``category_views.py`` en suivant votre convention
actuelle (un fichier de vues par entité). ``AdminRequiredMixin`` est
réimporté depuis ``category_views`` plutôt que redéfini : adaptez cet
import si vous renommez ce fichier.
"""
from __future__ import annotations

from typing import Any, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Q, QuerySet, Sum
from django.db.models.functions import Coalesce
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils.translation import gettext_lazy as _
from django.views import View
from django.views.generic import CreateView, ListView, TemplateView, UpdateView

from products.forms import ProduitForm, ProduitImportForm
from products.models import ImportProduits, Produit
from products.services import get_categories_actives
from products.tasks import importer_produits
from products.views.category_views import AdminRequiredMixin


class ProduitListView(LoginRequiredMixin, ListView):
    """Liste paginée des produits, avec recherche et filtre catégorie (EF-3, EF-13.3).

    Optimisations :
      - ``select_related("categorie")`` : évite une requête par ligne pour
        afficher le nom de la catégorie (EF-13.1).
      - ``annotate`` (Count/Sum) pour le nombre d'agences et le stock
        total plutôt qu'une requête ProduitAgence par produit affiché.
      - filtre catégorie peuplé depuis ``get_categories_actives()``, mis
        en cache (products.services), pas une requête à chaque affichage.
      - pagination (``paginate_by``).
    """

    model = Produit
    template_name = "dashboard/pages/admin/products/product/produit_list.html"
    context_object_name = "produits"
    paginate_by = 20

    def get_queryset(self) -> QuerySet[Produit]:
        qs = (
            Produit.objects.select_related("categorie")
            .annotate(
                nombre_agences=Count(
                    "prix_stocks", filter=Q(prix_stocks__actif=True), distinct=True
                ),
                stock_total=Coalesce(Sum("prix_stocks__stock_quantite"), 0),
            )
            .order_by("nom")
        )

        q = self.request.GET.get("q", "").strip()
        if q:
            qs = qs.filter(Q(nom__icontains=q) | Q(sku__icontains=q))

        categorie_id = self.request.GET.get("categorie", "").strip()
        if categorie_id:
            qs = qs.filter(categorie_id=categorie_id)

        return qs

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context["q"] = self.request.GET.get("q", "")
        context["categorie_id"] = self.request.GET.get("categorie", "")
        context["categories"] = get_categories_actives()
        return context


class ProduitCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    """Création d'un produit (EF-3.1)."""

    model = Produit
    form_class = ProduitForm
    template_name = "dashboard/pages/admin/products/product/produit_form.html"
    success_url = reverse_lazy("products:produit_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _("Produit « %(nom)s » créé.") % {"nom": self.object.nom})
        return response


class ProduitUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    """Modification d'un produit existant (EF-3.1).

    Réutilise ``ProduitForm`` et le même template que la création — seuls
    le titre et le texte du bouton changent selon que ``object`` est
    défini (édition) ou non (création), voir ``produit_form.html``.
    """

    model = Produit
    form_class = ProduitForm
    template_name = "dashboard/pages/admin/products/product/produit_form.html"
    success_url = reverse_lazy("products:produit_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _("Produit « %(nom)s » mis à jour.") % {"nom": self.object.nom})
        return response


class ProduitImportView(LoginRequiredMixin, AdminRequiredMixin, TemplateView):
    """Upload d'un fichier Excel de produits, traité en arrière-plan (EF-3.1).

    Même principe que l'import de catégories : la requête ne fait que
    sauvegarder le fichier et déclencher la tâche Celery
    (``importer_produits.delay``), donc elle ne bloque jamais, même pour
    un fichier de plusieurs milliers de lignes.
    """

    template_name = "dashboard/pages/admin/products/product/produit_import.html"

    def get_context_data(self, **kwargs: Any) -> Dict[str, Any]:
        context = super().get_context_data(**kwargs)
        context.setdefault("form", ProduitImportForm())
        context["imports"] = (
            ImportProduits.objects.select_related("importe_par").order_by("-created_at")[:10]
        )
        return context

    def post(self, request, *args, **kwargs):
        form = ProduitImportForm(request.POST, request.FILES)
        if form.is_valid():
            import_obj = form.save(commit=False)
            if request.user.is_authenticated:
                import_obj.importe_par = request.user
            import_obj.save()
            importer_produits.delay(str(import_obj.pk))
            messages.success(
                request,
                _("Import lancé. Le rapport apparaîtra ci-dessous une fois le traitement terminé."),
            )
            return redirect("products:produit_import")

        context = self.get_context_data(form=form)
        return self.render_to_response(context)


class ProduitImportTemplateView(LoginRequiredMixin, AdminRequiredMixin, View):
    """Génère et sert le modèle Excel attendu par l'import de produits (EF-3.1)."""

    FILENAME = "modele_import_produits.xlsx"
    HEADERS = ["nom", "sku", "categorie", "prix_achat", "prix_vente_defaut", "unite", "description", "actif"]
    EXEMPLES = [
        ["Riz parfumé 5kg", "RIZ-5KG", "Alimentation", 3500, 4200, "sac", "Riz parfumé importé", 1],
        ["Coca-Cola 1.5L", "COCA-150", "Boissons gazeuses", 600, 900, "bouteille", "", 1],
        ["Savon Marseille", "SAV-MRS", "Hygiène (désactivée)", 250, 400, "unité", "", 0],
    ]

    def get(self, request: HttpRequest, *args: Any, **kwargs: Any) -> HttpResponse:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Produits"

        sheet.append(self.HEADERS)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        for col_idx in range(1, len(self.HEADERS) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill

        for row in self.EXEMPLES:
            sheet.append(row)

        for col_idx, width in enumerate([25, 15, 22, 12, 18, 12, 30, 8], start=1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{self.FILENAME}"'
        workbook.save(response)
        return response