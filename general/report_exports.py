"""Génération des exports Excel et PDF des rapports (EF-11, EF-12).

Deux formats génériques, réutilisés par les rapports d'inventaire et de
ventes : `exporter_excel()` (openpyxl) et `exporter_pdf()` (WeasyPrint,
même approche que `sales/receipts.py`). Les deux bibliothèques sont
importées LOCALEMENT dans chaque fonction, pas en haut du module : si
l'une d'elles n'est pas encore installée, le reste du projet (y compris
l'autre export) continue de fonctionner normalement.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Iterable, Sequence

from django.http import HttpResponse
from django.template.loader import render_to_string


def exporter_excel(
    *,
    nom_fichier: str,
    entetes: Sequence[str],
    lignes: Iterable[Sequence[Any]],
    titre: str = "",
) -> HttpResponse:
    """Construit un classeur Excel à une feuille et le renvoie en pièce jointe."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Rapport"

    ligne_entetes = 1
    if titre:
        feuille.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(entetes), 1))
        cellule_titre = feuille.cell(row=1, column=1, value=titre)
        cellule_titre.font = Font(bold=True, size=13)
        ligne_entetes = 3

    remplissage_entete = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    police_entete = Font(bold=True, color="FFFFFF")
    for colonne, libelle in enumerate(entetes, start=1):
        cellule = feuille.cell(row=ligne_entetes, column=colonne, value=libelle)
        cellule.font = police_entete
        cellule.fill = remplissage_entete

    nombre_lignes = 0
    for decalage, ligne in enumerate(lignes, start=1):
        nombre_lignes = decalage
        for colonne, valeur in enumerate(ligne, start=1):
            feuille.cell(row=ligne_entetes + decalage, column=colonne, value=valeur)

    if nombre_lignes == 0:
        feuille.cell(row=ligne_entetes + 1, column=1, value="Aucune donnée pour ces filtres.")

    for colonne in range(1, len(entetes) + 1):
        feuille.column_dimensions[get_column_letter(colonne)].width = 22

    tampon = BytesIO()
    classeur.save(tampon)
    tampon.seek(0)

    reponse = HttpResponse(
        tampon.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    reponse["Content-Disposition"] = f'attachment; filename="{nom_fichier}"'
    return reponse


def exporter_pdf(*, nom_fichier: str, nom_gabarit: str, contexte: dict) -> HttpResponse:
    """Rend un gabarit HTML autonome (pas de dashboard/base.html) en PDF."""
    from weasyprint import HTML

    html = render_to_string(nom_gabarit, contexte)
    pdf_bytes = HTML(string=html).write_pdf()

    reponse = HttpResponse(pdf_bytes, content_type="application/pdf")
    reponse["Content-Disposition"] = f'attachment; filename="{nom_fichier}"'
    return reponse